#!/usr/bin/python
# -*- coding: utf-8 -*- 
"""This scripts contains function that help dealing with language file for android

    Created by vietnh
    
    Usage:
        - Init excel sheet:
            ./migrate_language.py dailyreport-adr -i

        - Update from excel file to language files:
            ./migrate_language.py dailyreport-adr

        - Add new strings to existing excel file:
            ./migrate_language.py dailyreport-adr -r
        
        
"""

import re, sys
from openpyxl import load_workbook
from openpyxl.styles import Font, colors

FILE_EN = '../' + sys.argv[1] + '/src/main/res/values/strings.xml'
FILE_JA = '../' + sys.argv[1] + '/src/main/res/values-ja/strings.xml'
FILE_VI = '../' + sys.argv[1] + '/src/main/res/values-vi/strings.xml'
FILE_LANGUAGE = '../language.xlsx'

# from android string resource files, export to excel file
# with key from en
def init_excel_file(file_en, file_ja, file_vi, file_out):
    wb = load_workbook(file_out)
    ws = wb.create_sheet(title=sys.argv[1])

    fen = open(file_en, 'r')
    fja = open(file_ja, 'r')
    fvi = open(file_vi, 'r')

    lines_ja = fja.readlines()
    lines_vi = fvi.readlines()

    font = Font(color=colors.RED)
    ws.cell(row = 1, column = 1).value = "Key"
    ws.cell(row = 1, column = 1).font = font
    ws.cell(row = 1, column = 2).value = "English"
    ws.cell(row = 1, column = 2).font = font
    ws.cell(row = 1, column = 3).value = "Japanese"
    ws.cell(row = 1, column = 3).font = font
    ws.cell(row = 1, column = 4).value = "Vietnamese"
    ws.cell(row = 1, column = 4).font = font


    row = 2
    col = 1
    for line_i in fen:
        # <string name="app_name">カレンダー</string>
        # <resources>
        # <!--Notifications-->
        if  re.search('<string', line_i): #if it's not comment line
            # print "line_i == " + line_i
            key = line_i[line_i.index('name="') + 6 : line_i.index('">')]
            print "Inserting", key
            en_value = line_i[line_i.index('">') + 2 : line_i.index('</string>')]
            print "en", en_value
            ws.cell(row = row, column = col).value = key
            ws.cell(row = row, column = col + 1).value = en_value
            for japanese_string in lines_ja:
                if  re.search('<string', japanese_string): #if it's not comment line
                    ja_key = japanese_string[japanese_string.index('name="') + 6 : japanese_string.index('">')]
                    if key == ja_key:
                        ja_value = japanese_string[japanese_string.index('">') + 2 : japanese_string.index('</string>')]
                        print "ja", ja_value
                        ws.cell(row = row, column = col + 2).value = ja_value

            for vi_string in lines_vi:
                if  re.search('<string', vi_string): #if it's not comment line
                    vi_key = vi_string[vi_string.index('name="') + 6 : vi_string.index('">')]
                    if key == vi_key:
                        vi_value = vi_string[vi_string.index('">') + 2 : vi_string.index('</string>')]
                        print "vi", vi_value
                        ws.cell(row = row, column = col + 3).value = vi_value

            row = row + 1

    fen.close()
    fja.close()
    fvi.close()

    wb.save(FILE_LANGUAGE)
#end of init_excel_file

def append_to_row(rows, key, value, language):
    existed = False
    for row in rows:
        if row['key'] == key:
            row[language] = value
            existed = True
    if not existed:
        row = {'key': key}
        row[language] = value
        rows.append(row)

# get data from android language files, update existing excel file
def add_new_keys_to_excel_file(file_en, file_ja, file_vi, file_out):
    wb = load_workbook(file_out)
    ws = wb[sys.argv[1]]

    fen = open(file_en, 'r')
    fja = open(file_ja, 'r')
    fvi = open(file_vi, 'r')

    rows = []

    for line_i in fen:
        if  re.search('<string', line_i): #if it's not comment line
            key = line_i[line_i.index('name="') + 6 : line_i.index('">')]
            if key not in map(lambda x: x.value, ws['A']):
                value = line_i[line_i.index('">') + 2 : line_i.index('</string>')]
                append_to_row(rows, key, value, 'en')

    for line_i in fja:
        if  re.search('<string', line_i): #if it's not comment line
            key = line_i[line_i.index('name="') + 6 : line_i.index('">')]
            if key not in map(lambda x: x.value, ws['A']):
                value = line_i[line_i.index('">') + 2 : line_i.index('</string>')]
                append_to_row(rows, key, value, 'ja')

    for line_i in fvi:
        if  re.search('<string', line_i): #if it's not comment line
            key = line_i[line_i.index('name="') + 6 : line_i.index('">')]
            if key not in map(lambda x: x.value, ws['A']):
                value = line_i[line_i.index('">') + 2 : line_i.index('</string>')]
                append_to_row(rows, key, value, 'vi')

    next_row_in_excel = len(ws['A']) + 1

    # print "next_row_in_excel" , next_row_in_excel
    # print rows

    font = Font(color=colors.RED)

    for row in rows:
        value_en, value_ja, value_vi = "", "", ""
        if 'en' in row:
            value_en = row['en']
        if 'ja' in row:
            value_ja = row['ja']
        if 'vi' in row:
            value_vi = row['vi']

        ws.cell(row = next_row_in_excel, column = 1).value = row['key']
        ws.cell(row = next_row_in_excel, column = 1).font = font
        ws.cell(row = next_row_in_excel, column = 2).value = value_en
        ws.cell(row = next_row_in_excel, column = 2).font = font
        ws.cell(row = next_row_in_excel, column = 3).value = value_ja
        ws.cell(row = next_row_in_excel, column = 3).font = font
        ws.cell(row = next_row_in_excel, column = 4).value = value_vi
        ws.cell(row = next_row_in_excel, column = 4).font = font

        next_row_in_excel += 1

    print "Added: ", rows

    fen.close()
    fja.close()
    fvi.close()

    wb.save(FILE_LANGUAGE)

def rewrite_to_file(f_out, data):
    f_out.seek(0)
    f_out.truncate()
    for li in data:
        f_out.write(li.encode('utf-8'))

# Import language from excel files to android language files
def migrate_languages():
    wb = load_workbook(filename = FILE_LANGUAGE, read_only=True)
    ws = wb[sys.argv[1]]

    fen = open(FILE_EN, 'r+')
    fja = open(FILE_JA, 'r+')
    fvi = open(FILE_VI, 'r+')

    copy_en = ["<resources>\n"]
    copy_ja = ["<resources>\n"]
    copy_vi = ["<resources>\n"]

    for idx, row in enumerate(ws.rows):
        if idx >= 1: #ignore first row
            key = row[0].value
            en_value = row[1].value
            if en_value is None:
                en_value = ""
            ja_value = row[2].value
            if ja_value is None:
                ja_value = ""
            vi_value = row[3].value
            if vi_value is None:
                vi_value = ""

            if key is not None:
                print "Migrating", key, '|', en_value, '|', ja_value, '|', vi_value
                copy_en.append('    <string name="' + key + '">' + en_value + '</string>' + "\n")
                copy_ja.append('    <string name="' + key + '">' + ja_value + '</string>' + "\n")
                copy_vi.append('    <string name="' + key + '">' + vi_value + '</string>' + "\n")


    copy_en.append("</resources>")
    copy_ja.append("</resources>")
    copy_vi.append("</resources>")

    rewrite_to_file(fen, copy_en)
    rewrite_to_file(fja, copy_ja)
    rewrite_to_file(fvi, copy_vi)

    fen.close()
    fja.close()
    fvi.close()


if __name__ == '__main__':
    print sys.argv
    # migrate_languages()
    if len(sys.argv) >= 3 and sys.argv[2] == '-i':
        init_excel_file(FILE_EN, FILE_JA, FILE_VI, FILE_LANGUAGE)
    elif len(sys.argv) >= 3 and sys.argv[2] == '-r':
        add_new_keys_to_excel_file(FILE_EN, FILE_JA, FILE_VI, FILE_LANGUAGE)
    else:
        migrate_languages()
